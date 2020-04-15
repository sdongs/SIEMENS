# -*- coding: utf-8 -*-
"""
Created on Thu Mar 26 16:29:44 2020

@author: sds
"""
from selenium import webdriver
import time
from bs4 import BeautifulSoup
import re
import openpyxl

filename='台账汇总.xlsx'
wb = openpyxl.load_workbook(filename)
#worksheet = workbook.add_worksheet(u'sheet1')#在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
sh = wb['Sheet1']
print(sh.max_row)

rstr = r"[\/\\\:\*\?\"\<\>\|&@%$#]"

for i in range(520,sh.max_row+1):
    d = sh.cell(row=i, column=4).value.replace(' ', '')
    print(d)    
    print(i)
    browser = webdriver.Firefox()
    browser.get("https://mall.industry.siemens.com/mall/zh/cn/Catalog/Search?searchTerm=%s&tab=Product" %d)
    time.sleep(6)
    content = browser.page_source
    #print(content)

    browser.quit() # 退出浏览器
    d=re.sub(rstr, "",d)
    f=open("%s.txt" %d,'w', encoding='UTF-8')
    f.write(content)
    soup = BeautifulSoup(content,'html.parser',from_encoding='utf-8')
    #print(soup)

    ceshi =soup.find_all('div',id="FuzzySearchInfo")
    print(ceshi[0].attrs.get('style'))
    if ceshi[0].attrs.get('style')=='font-size: 14px;':
        res1='订货号未查到'
    else:
        res1='订货号已查到'
    notfound=re.findall('抱歉，您的搜索请求',content)
    print(notfound)
    if len(notfound)!=0:
        res1='订货号未查到'
    print(res1)
    sh.cell(row=i,column=5).value=res1
    #display: none; font-size: 14px
    #font-size: 14px;

    notavailable=soup.find_all('div',class_="blueLineMessage")
    if len(notavailable)!=0:
        print(notavailable[0].text.replace(' ', '').replace('\n', '').replace('\r', ''))
        res2=notavailable[0].text.replace(' ', '').replace('\n', '').replace('\r', '')
        res21 = re.findall(r'后继产品：(.*?)比较产品', res2)
        sh.cell(row=i,column=6).value=res2
        if len(res21)!=0:
            sh.cell(row=i,column=7).value=res21[0]

    products =soup.find_all('td',class_="ProductName")
    j=8
    for product in products:
        a = product.find_all('div')
        res3=a[0].text.replace(' ', '').replace('\n', '').replace('\r', '')
        res4=a[1].text.replace(' ', '').replace('\n', '').replace('\r', '')
        sh.cell(row=i,column=j).value=res3
        sh.cell(row=i,column=j+1).value=res4
        j=j+2
    wb.save(filename)
    





